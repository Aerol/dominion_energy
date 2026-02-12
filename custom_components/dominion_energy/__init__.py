"""The Dominion Energy Virginia integration."""
from __future__ import annotations

import logging
from datetime import timedelta

from homeassistant.config_entries import ConfigEntry
from homeassistant.const import Platform, CONF_USERNAME, CONF_PASSWORD
from homeassistant.core import HomeAssistant
from homeassistant.helpers.aiohttp_client import async_get_clientsession
from homeassistant.helpers.update_coordinator import DataUpdateCoordinator, UpdateFailed

from .api import DominionEnergyAPI, DominionEnergyAPIError, TokenExpiredError
from .const import DOMAIN, CONF_MANUAL_TOKEN

_LOGGER = logging.getLogger(__name__)

PLATFORMS: list[Platform] = [Platform.SENSOR]


async def async_setup_entry(hass: HomeAssistant, entry: ConfigEntry) -> bool:
    """Set up Dominion Energy Virginia from a config entry."""
    session = async_get_clientsession(hass)
    auth_method = entry.data.get("auth_method", "manual_token")

    if auth_method == "automatic":
        # Automatic authentication with tokens
        username = entry.data[CONF_USERNAME]
        password = entry.data[CONF_PASSWORD]
        access_token = entry.data.get("access_token")
        refresh_token = entry.data.get("refresh_token")

        api = DominionEnergyAPI(
            username=username,
            password=password,
            session=session,
            access_token=access_token,
            refresh_token=refresh_token,
        )

        # Import cookies if available (for 2FA bypass)
        cookies = entry.data.get("cookies", {})
        if cookies:
            api.import_cookies(cookies)

        # Set account numbers from config
        api._account_number = entry.data.get("account_number")
        api._customer_number = entry.data.get("customer_number")
        api._meter_number = entry.data.get("meter_number")
        
        _LOGGER.error("DEBUG: Config entry data keys: %s", list(entry.data.keys()))
        _LOGGER.error("DEBUG: Account number from config: %s", api._account_number)
        _LOGGER.error("DEBUG: Customer number from config: %s", api._customer_number)
        _LOGGER.error("DEBUG: Meter number from config: %s", api._meter_number)
        
        if not api._account_number or not api._meter_number:
            _LOGGER.error("Missing account/meter numbers in config entry! Please reconfigure integration.")
            return False

        # Ensure token is valid (will auto-refresh if needed)
        try:
            await api.async_ensure_valid_token()
            _LOGGER.info("Token validated successfully")

            # If token was refreshed, update config entry
            if api._access_token != access_token:
                hass.config_entries.async_update_entry(
                    entry,
                    data={
                        **entry.data,
                        "access_token": api._access_token,
                        "refresh_token": api._refresh_token,
                    },
                )
                _LOGGER.info("Tokens refreshed and saved")

        except TokenExpiredError:
            # Refresh token expired - try to re-authenticate with stored credentials
            _LOGGER.warning("Refresh token expired, attempting automatic re-authentication")
            
            try:
                # Re-authenticate using stored credentials and cookies
                _LOGGER.info("Attempting login with stored credentials")
                result = await api.async_login_with_credentials(username, password)
                
                if result.get("tfa_required"):
                    # 2FA required but we can't prompt user during setup
                    # Mark integration for reconfiguration
                    _LOGGER.error("2FA required for re-authentication - please reconfigure integration")
                    hass.components.persistent_notification.create(
                        title="Dominion Energy - Re-authentication Required",
                        message="Your session has expired and 2FA is required. Please delete and re-add the integration.",
                        notification_id="dominion_energy_reauth"
                    )
                    return False
                else:
                    # Login successful, got new tokens
                    tokens = await api._async_exchange_for_dominion_tokens()
                    
                    # Save new tokens
                    hass.config_entries.async_update_entry(
                        entry,
                        data={
                            **entry.data,
                            "access_token": tokens.access_token,
                            "refresh_token": tokens.refresh_token,
                            "cookies": api.export_cookies(),
                        },
                    )
                    _LOGGER.info("Re-authentication successful, tokens saved")
                    
            except Exception as e:
                _LOGGER.error("Re-authentication failed: %s", e)
                hass.components.persistent_notification.create(
                    title="Dominion Energy - Re-authentication Failed",
                    message=f"Automatic re-authentication failed: {e}. Please delete and re-add the integration.",
                    notification_id="dominion_energy_reauth"
                )
                return False
                
        except DominionEnergyAPIError as e:
            _LOGGER.error("Failed to validate token: %s", e)
            return False

        # Get account info (optional - integration can work without it)
        try:
            await api.async_get_account_info()
            _LOGGER.info("Account info retrieved successfully")
        except DominionEnergyAPIError as e:
            _LOGGER.warning("Could not get account info (not critical): %s", e)
            # Continue anyway - tokens work fine

    else:
        # Manual token method
        manual_token = entry.data[CONF_MANUAL_TOKEN]
        account_number = entry.data["account_number"]
        customer_number = entry.data["customer_number"]
        meter_number = entry.data["meter_number"]

        api = DominionEnergyAPI(
            username="",
            password="",
            session=session,
            access_token=manual_token,
            refresh_token=None,
        )

        api._account_number = account_number
        api._customer_number = customer_number
        api._meter_number = meter_number

    # Create data update coordinator
    coordinator = DominionEnergyDataUpdateCoordinator(hass, api)

    # Fetch initial data so sensors have something to display
    await coordinator.async_config_entry_first_refresh()

    hass.data.setdefault(DOMAIN, {})
    hass.data[DOMAIN][entry.entry_id] = {
        "api": api,
        "coordinator": coordinator,
    }

    await hass.config_entries.async_forward_entry_setups(entry, PLATFORMS)

    return True


async def async_unload_entry(hass: HomeAssistant, entry: ConfigEntry) -> bool:
    """Unload a config entry."""
    if unload_ok := await hass.config_entries.async_unload_platforms(entry, PLATFORMS):
        hass.data[DOMAIN].pop(entry.entry_id)

    return unload_ok


class DominionEnergyDataUpdateCoordinator(DataUpdateCoordinator):
    """Class to manage fetching Dominion Energy data."""

    def __init__(self, hass: HomeAssistant, api: DominionEnergyAPI) -> None:
        """Initialize coordinator."""
        super().__init__(
            hass,
            _LOGGER,
            name=DOMAIN,
            update_interval=timedelta(minutes=25),  # Refresh before 30min token expiry
        )
        self.api = api

    def _parse_excel_data(self, excel_bytes: bytes) -> dict:
        """Parse Excel data and return totals."""
        try:
            from openpyxl import load_workbook
            from io import BytesIO
            
            wb = load_workbook(BytesIO(excel_bytes))
            ws = wb.active
            
            total_usage = 0
            last_value = 0
            last_time = None
            reading_count = 0
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 4:
                    continue
                
                date_str = row[2]
                
                for col_idx in range(3, min(len(row), 51)):
                    value = row[col_idx]
                    if value is not None and value != '':
                        try:
                            usage = float(value)
                            total_usage += usage
                            last_value = usage
                            last_time = f"{date_str} {ws.cell(1, col_idx + 1).value}"
                            reading_count += 1
                        except (ValueError, TypeError):
                            pass
            
            return {
                "total": total_usage,
                "last_value": last_value,
                "last_time": last_time,
                "count": reading_count
            }
        except Exception as e:
            _LOGGER.error("Error parsing Excel data: %s", e, exc_info=True)
            return {"total": 0, "last_value": 0, "last_time": None, "count": 0}

    def _parse_green_button_data(self, xml_bytes: bytes) -> float:
        """Parse Green Button XML and return total kWh."""
        try:
            import xml.etree.ElementTree as ET
            
            root = ET.fromstring(xml_bytes)
            
            ET.register_namespace('espi', 'http://naesb.org/espi')
            ns = {'espi': 'http://naesb.org/espi'}
            
            # Get powerOfTenMultiplier
            power_multiplier = 0
            reading_type = root.find('.//espi:ReadingType', ns)
            if reading_type is not None:
                power_elem = reading_type.find('espi:powerOfTenMultiplier', ns)
                if power_elem is not None and power_elem.text:
                    power_multiplier = int(power_elem.text)
            
            # Find all IntervalReading elements
            readings = root.findall('.//espi:IntervalReading', ns)
            
            total = 0
            for reading in readings:
                value_elem = reading.find('espi:value', ns)
                if value_elem is not None and value_elem.text:
                    try:
                        raw_value = int(value_elem.text)
                        wh = raw_value * (10 ** power_multiplier)
                        kwh = wh / 1000.0
                        total += kwh
                    except (ValueError, TypeError):
                        pass
            
            return total
        except Exception as e:
            _LOGGER.error("Error parsing Green Button XML: %s", e, exc_info=True)
            return 0

    async def _async_update_data(self):
        """Fetch data from API."""
        try:
            # Ensure token is valid (auto-refresh if needed)
            if self.api.has_tokens:
                await self.api.async_ensure_valid_token()

            # Fetch BOTH Excel and Green Button data for comparison
            _LOGGER.error("DEBUG: Fetching usage data with account=%s, meter=%s", 
                         self.api._account_number, self.api._meter_number)
            
            from datetime import datetime, timedelta
            
            # Define date ranges
            today = datetime.now().date()
            yesterday = today - timedelta(days=1)
            first_of_month = today.replace(day=1)
            
            # Get actual billing period from Dominion API
            _LOGGER.error("DEBUG: Fetching billing information...")
            billing_info = await self.api.async_get_billing_info()
            
            billing_start = None
            billing_end = None
            amount_due = None
            current_charges = None
            
            if billing_info and billing_info.get("status", {}).get("code") == 200:
                data = billing_info.get("data", {})
                results = data.get("zBillInvHeadtoItemNav", {}).get("results", [])
                
                if results:
                    latest_bill = results[0]
                    
                    # Parse billing dates
                    bill_start_str = latest_bill.get("billPdStart", "")
                    bill_end_str = latest_bill.get("billPdEnd", "")
                    
                    if bill_start_str and bill_end_str:
                        try:
                            # Parse format: "12/16/2025 00:00:00"
                            billing_start = datetime.strptime(bill_start_str, "%m/%d/%Y %H:%M:%S").date()
                            billing_end = datetime.strptime(bill_end_str, "%m/%d/%Y %H:%M:%S").date()
                            
                            # Get amounts
                            amount_due = float(latest_bill.get("amountDue", "0"))
                            current_charges = float(latest_bill.get("totalCurrentCharges", "0"))
                            
                            _LOGGER.error("DEBUG: Actual billing period from API: %s to %s", 
                                         billing_start, billing_end)
                            _LOGGER.error("DEBUG: Amount due: $%.2f, Current charges: $%.2f", 
                                         amount_due, current_charges)
                        except (ValueError, AttributeError) as e:
                            _LOGGER.warning("Could not parse billing dates: %s", e)
            
            # If we got billing dates from API, use them
            # Otherwise fall back to assuming 17th of month
            if billing_start and billing_end:
                # Use the actual billing period from Dominion
                # But extend to today if we're past the bill end date
                if today > billing_end:
                    # We're in the new billing period, use bill_end as start
                    actual_billing_start = billing_end + timedelta(days=1)
                else:
                    # We're still in the current billing period
                    actual_billing_start = billing_start
            else:
                # Fallback: Dominion billing period (typically 17th to 17th)
                billing_day = 17
                
                # Calculate billing period start date
                if today.day >= billing_day:
                    actual_billing_start = today.replace(day=billing_day)
                else:
                    last_month = today.replace(day=1) - timedelta(days=1)
                    actual_billing_start = last_month.replace(day=billing_day)
            
            _LOGGER.error("DEBUG: Using billing start date: %s to %s", actual_billing_start, today)
            
            # Fetch YESTERDAY's data for daily usage (today's data isn't finalized yet)
            _LOGGER.error("DEBUG: === Fetching YESTERDAY's data (%s) for daily usage ===", yesterday)
            daily_excel = await self.api.async_get_usage(
                start_date=datetime.combine(yesterday, datetime.min.time()),
                end_date=datetime.combine(yesterday, datetime.max.time())
            )
            
            try:
                daily_green_button = await self.api.async_get_green_button_data(
                    start_date=datetime.combine(yesterday, datetime.min.time()),
                    end_date=datetime.combine(yesterday, datetime.max.time())
                )
            except Exception as e:
                _LOGGER.error("DEBUG: Failed to fetch daily Green Button data: %s", e)
                daily_green_button = None
            
            # Fetch MONTH's data for monthly usage
            _LOGGER.error("DEBUG: === Fetching CALENDAR MONTH data (%s to %s) ===", first_of_month, today)
            monthly_excel = await self.api.async_get_usage(
                start_date=datetime.combine(first_of_month, datetime.min.time()),
                end_date=datetime.combine(today, datetime.max.time())
            )
            
            try:
                monthly_green_button = await self.api.async_get_green_button_data(
                    start_date=datetime.combine(first_of_month, datetime.min.time()),
                    end_date=datetime.combine(today, datetime.max.time())
                )
            except Exception as e:
                _LOGGER.error("DEBUG: Failed to fetch monthly Green Button data: %s", e)
                monthly_green_button = None
            
            # Fetch BILLING PERIOD data (using actual dates from Dominion API)
            _LOGGER.error("DEBUG: === Fetching BILLING PERIOD data (%s to %s) ===", actual_billing_start, today)
            billing_excel = await self.api.async_get_usage(
                start_date=datetime.combine(actual_billing_start, datetime.min.time()),
                end_date=datetime.combine(today, datetime.max.time())
            )
            
            try:
                billing_green_button = await self.api.async_get_green_button_data(
                    start_date=datetime.combine(actual_billing_start, datetime.min.time()),
                    end_date=datetime.combine(today, datetime.max.time())
                )
            except Exception as e:
                _LOGGER.error("DEBUG: Failed to fetch billing Green Button data: %s", e)
                billing_green_button = None

            # Parse the response into sensor data
            parsed_data = {
                "account_number": self.api._account_number,
                "meter_number": self.api._meter_number,
                "last_hour_usage": 0,
                "last_hour_reading_time": None,
                "daily_usage": 0,
                "monthly_usage": 0,
                "billing_usage": 0,
                "amount_due": amount_due if amount_due else 0,
                "current_charges": current_charges if current_charges else 0,
                "estimated_cost": 0,
                "daily_excel": 0,
                "daily_green_button": 0,
                "monthly_excel": 0,
                "monthly_green_button": 0,
                "billing_excel": 0,
                "billing_green_button": 0,
            }
            
            # Parse DAILY data
            daily_excel_total = 0
            daily_gb_total = 0
            last_value = 0
            last_time = None
            
            if daily_excel and isinstance(daily_excel, dict) and "raw_excel" in daily_excel:
                _LOGGER.error("DEBUG: === Processing DAILY Excel ===")
                excel_result = self._parse_excel_data(daily_excel["raw_excel"])
                daily_excel_total = excel_result["total"]
                last_value = excel_result["last_value"]
                last_time = excel_result["last_time"]
                _LOGGER.error("DEBUG: Daily Excel: %.2f kWh", daily_excel_total)
            
            if daily_green_button and isinstance(daily_green_button, dict) and "raw_xml" in daily_green_button:
                _LOGGER.error("DEBUG: === Processing DAILY Green Button ===")
                daily_gb_total = self._parse_green_button_data(daily_green_button["raw_xml"])
                _LOGGER.error("DEBUG: Daily Green Button: %.2f kWh", daily_gb_total)
            
            # Use Green Button for daily if available, otherwise Excel
            parsed_data["daily_usage"] = daily_gb_total if daily_gb_total > 0 else daily_excel_total
            parsed_data["daily_excel"] = daily_excel_total
            parsed_data["daily_green_button"] = daily_gb_total
            parsed_data["last_hour_usage"] = last_value
            parsed_data["last_hour_reading_time"] = last_time
            
            # Parse MONTHLY data
            monthly_excel_total = 0
            monthly_gb_total = 0
            
            if monthly_excel and isinstance(monthly_excel, dict) and "raw_excel" in monthly_excel:
                _LOGGER.error("DEBUG: === Processing MONTHLY Excel ===")
                excel_result = self._parse_excel_data(monthly_excel["raw_excel"])
                monthly_excel_total = excel_result["total"]
                _LOGGER.error("DEBUG: Monthly Excel: %.2f kWh", monthly_excel_total)
            
            if monthly_green_button and isinstance(monthly_green_button, dict) and "raw_xml" in monthly_green_button:
                _LOGGER.error("DEBUG: === Processing MONTHLY Green Button ===")
                monthly_gb_total = self._parse_green_button_data(monthly_green_button["raw_xml"])
                _LOGGER.error("DEBUG: Monthly Green Button: %.2f kWh", monthly_gb_total)
            
            # Use Green Button for monthly if available, otherwise Excel
            parsed_data["monthly_usage"] = monthly_gb_total if monthly_gb_total > 0 else monthly_excel_total
            parsed_data["monthly_excel"] = monthly_excel_total
            parsed_data["monthly_green_button"] = monthly_gb_total
            
            # Parse BILLING PERIOD data
            billing_excel_total = 0
            billing_gb_total = 0
            
            if billing_excel and isinstance(billing_excel, dict) and "raw_excel" in billing_excel:
                _LOGGER.error("DEBUG: === Processing BILLING PERIOD Excel ===")
                excel_result = self._parse_excel_data(billing_excel["raw_excel"])
                billing_excel_total = excel_result["total"]
                _LOGGER.error("DEBUG: Billing Period Excel: %.2f kWh", billing_excel_total)
            
            if billing_green_button and isinstance(billing_green_button, dict) and "raw_xml" in billing_green_button:
                _LOGGER.error("DEBUG: === Processing BILLING PERIOD Green Button ===")
                billing_gb_total = self._parse_green_button_data(billing_green_button["raw_xml"])
                _LOGGER.error("DEBUG: Billing Period Green Button: %.2f kWh", billing_gb_total)
            
            # Use Green Button for billing if available, otherwise Excel
            parsed_data["billing_usage"] = billing_gb_total if billing_gb_total > 0 else billing_excel_total
            parsed_data["billing_excel"] = billing_excel_total
            parsed_data["billing_green_button"] = billing_gb_total
            
            # Use actual bill amount if available, otherwise estimate
            if amount_due and amount_due > 0:
                parsed_data["estimated_cost"] = amount_due
                _LOGGER.info("Using actual bill amount: $%.2f", amount_due)
            elif current_charges and current_charges > 0:
                parsed_data["estimated_cost"] = current_charges
                _LOGGER.info("Using current charges: $%.2f", current_charges)
            else:
                # Fallback to estimate based on usage
                parsed_data["estimated_cost"] = round(parsed_data["billing_usage"] * 0.12, 2)
                _LOGGER.info("Estimating cost based on usage: $%.2f", parsed_data["estimated_cost"])
            
            # Log comparison
            _LOGGER.error("DEBUG: === FINAL RESULTS ===")
            _LOGGER.error("DEBUG: Daily:   %.2f kWh (Excel: %.2f, GB: %.2f)", 
                         parsed_data["daily_usage"], daily_excel_total, daily_gb_total)
            _LOGGER.error("DEBUG: Monthly: %.2f kWh (Excel: %.2f, GB: %.2f)", 
                         parsed_data["monthly_usage"], monthly_excel_total, monthly_gb_total)
            _LOGGER.error("DEBUG: Billing: %.2f kWh (Excel: %.2f, GB: %.2f)", 
                         parsed_data["billing_usage"], billing_excel_total, billing_gb_total)

            return parsed_data

        except TokenExpiredError as err:
            _LOGGER.exception("Token expired error: %s", err)
            raise UpdateFailed(f"Token expired: {err}") from err
        except DominionEnergyAPIError as err:
            _LOGGER.exception("API error: %s", err)
            raise UpdateFailed(f"Error communicating with API: {err}") from err
        except Exception as err:
            _LOGGER.exception("Unexpected error parsing usage data: %s", err)
            raise UpdateFailed(f"Error parsing data: {err}") from err
